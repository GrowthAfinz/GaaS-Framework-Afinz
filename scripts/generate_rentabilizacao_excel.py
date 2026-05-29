#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Gerador de Relatório Rentabilização Copa 2026
Lê rentabilizacao_activities do Supabase e gera Excel diarizado.

Uso:
    python generate_rentabilizacao_excel.py [--start YYYY-MM-DD] [--end YYYY-MM-DD] [--output arquivo.xlsx]

Padrão: mês corrente, saída em outputs/rentabilizacao_copa2026_<data>.xlsx
"""

import sys, os, argparse, unicodedata
from datetime import datetime, timedelta
from collections import defaultdict
from pathlib import Path

try:
    from supabase import create_client
except ImportError:
    sys.exit("Instale supabase: pip install supabase")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("Instale openpyxl: pip install openpyxl")

# ── Configuração ──────────────────────────────────────────────────────────────
SUPABASE_URL = 'https://mipiwxadnpwtcgfcedym.supabase.co'
SUPABASE_KEY = 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Im1pcGl3eGFkbnB3dGNnZmNlZHltIiwicm9sZSI6ImFub24iLCJpYXQiOjE3Njk0NjU1NDUsImV4cCI6MjA4NTA0MTU0NX0.kIPhFfqvcJJh2S4yS2PsopmSYsZfC7ZNausumJGtmrM'

DAY_NAMES_PT = ['Dom', 'Seg', 'Ter', 'Qua', 'Qui', 'Sex', 'Sab']

# ── Blocos por aba ────────────────────────────────────────────────────────────
# Cada bloco: (nome display, lista de tokens que identificam no journey name)
RNT_BLOCKS = [
    ('Ativação',     ['ATIVACAO', 'ATIVAÇÃO', 'DESBLOQUEIO', 'INCENTIVO', 'POS_TOMBAMENTO', 'WELCOME']),
    ('Reativação',   ['REATIVACAO', 'REATIVAÇÃO']),
    ('Pós-Emissão',  ['NOVOS', 'CARTONISTAS']),
]

SEG_BLOCKS = [
    # Mais específico primeiro — Carrinho antes de Mulher para evitar falso positivo
    ('Carrinho Seguro',  ['CARRINHO']),
    ('Seguro Mulher',    ['SEGURO_MULHER']),
    ('Seguro Residência',['RESIDENCIA24H', 'RESIDENCIA']),
]

# ── Canal → posição dentro de um bloco (0-indexed) ──────────────────────────
# Estrutura por bloco: WPP(sent, del, clk) | EMAIL(sent, del, abe, clk) | SMS(sent, del) | PUSH(sent, del)
# Total: 3 + 4 + 2 + 2 = 11 colunas por bloco
CANAL_COLS = {
    'WhatsApp': {'sent': 0, 'del': 1, 'clk': 2},
    'E-mail':   {'sent': 3, 'del': 4, 'abe': 5, 'clk': 6},
    'SMS':      {'sent': 7, 'del': 8},
    'Push':     {'sent': 9, 'del': 10},
}
BLOCK_WIDTH = 11  # colunas por bloco
FIXED_COLS  = 7   # Data, Dia, LP Tráfego, LP Optins, Invt. Mídia, Cliques Mídia, CPC

# Labels de sub-header linha 3 (canal) e linha 4 (métrica)
CANAL_SUBHEADER = [
    'wpp', None, None,
    'e-mail', None, None, None,
    'sms', None,
    'push', None,
]
METRIC_SUBHEADER = [
    'enviados', 'entregas', 'cliques',
    'enviados', 'entregas', 'abertura', 'cliques',
    'enviados', 'entregas',
    'enviados', 'entregas',
]

# ── Cores ─────────────────────────────────────────────────────────────────────
C = {
    'header_dark':  '0F172A',
    'header_teal':  '0891B2',
    'ativacao':     '1D4ED8',
    'reativacao':   '065F46',
    'posemissao':   '7C3AED',
    'seg_mulher':   '9D174D',
    'seg_reside':   '1E40AF',
    'seg_carrinho': '0369A1',
    'wpp_fill':     'D1FAE5', 'wpp_font': '064E3B',
    'email_fill':   'DBEAFE', 'email_font': '1E3A5F',
    'sms_fill':     'FEF3C7', 'sms_font': '78350F',
    'push_fill':    'FFE4E6', 'push_font': '881337',
    'total_fill':   'F0F9FF',
    'zebra':        'F8FAFC',
    'weekend':      'E2E8F0',
    'kpi_fill':     '0F172A',
}

BLOCK_COLORS = {
    'Ativação':      C['ativacao'],
    'Reativação':    C['reativacao'],
    'Pós-Emissão':   C['posemissao'],
    'Seguro Mulher':     C['seg_mulher'],
    'Seguro Residência': C['seg_reside'],
    'Carrinho Seguro':   C['seg_carrinho'],
}

CANAL_STYLE = {
    'wpp':    (C['wpp_fill'],   C['wpp_font']),
    'e-mail': (C['email_fill'], C['email_font']),
    'sms':    (C['sms_fill'],   C['sms_font']),
    'push':   (C['push_fill'],  C['push_font']),
}

def canal_fill(metric_col: int) -> tuple[str, str]:
    """Retorna (fill, font) pelo índice relativo dentro do bloco."""
    if metric_col < 3:   return CANAL_STYLE['wpp']
    if metric_col < 7:   return CANAL_STYLE['e-mail']
    if metric_col < 9:   return CANAL_STYLE['sms']
    return CANAL_STYLE['push']

# ── Helpers de estilo ─────────────────────────────────────────────────────────
thin = Side(style='thin', color='CBD5E1')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
thick_top = Side(style='medium', color='475569')

def cell(ws, row, col):
    return ws.cell(row=row, column=col)

def style(c, value=None, bold=False, fill=None, font_color='000000', align='center', size=9, italic=False):
    if value is not None:
        c.value = value
    c.font = Font(name='Calibri', bold=bold, size=size, color=f'FF{font_color}', italic=italic)
    if fill:
        c.fill = PatternFill('solid', start_color=f'FF{fill}')
    c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
    c.border = BORDER

# ── Classificação de jornada → bloco ─────────────────────────────────────────

def _norm(s: str) -> str:
    """Remove acentos e normaliza para uppercase."""
    return ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn').upper()

def classify_bloco(jornada: str, blocks: list[tuple]) -> str | None:
    # Envolve com _ para garantir que o token começa numa "palavra"
    # Usa _TOKEN (sem exigir _ depois) para aceitar RESIDENCIA24H etc.
    # Isso evita ATIVACAO matchear dentro de REATIVACAO
    # porque a string tem _REATIVACAO_ (não _ATIVACAO)
    j = f'_{_norm(jornada)}_'
    for bloco_name, tokens in blocks:
        for token in tokens:
            t = _norm(token)
            if f'_{t}' in j:
                return bloco_name
    return None

# ── Parse de data ─────────────────────────────────────────────────────────────
def parse_date(val) -> datetime | None:
    if not val:
        return None
    if isinstance(val, datetime):
        return val.replace(tzinfo=None)
    s = str(val)[:10]
    for fmt in ['%Y-%m-%d', '%d/%m/%Y']:
        try:
            return datetime.strptime(s, fmt)
        except:
            pass
    return None

def date_range(start: datetime, end: datetime) -> list[datetime]:
    dates = []
    d = start
    while d <= end:
        dates.append(d)
        d += timedelta(days=1)
    return dates

# ── Fetch Supabase ─────────────────────────────────────────────────────────────
def fetch_rows(start: datetime, end: datetime) -> list[dict]:
    client = create_client(SUPABASE_URL, SUPABASE_KEY)
    rows = []
    page = 1000
    for offset in range(0, 100_000, page):
        res = client.table('rentabilizacao_activities').select('*') \
            .gte('Data de Disparo', start.strftime('%Y-%m-%d')) \
            .lte('Data de Disparo', end.strftime('%Y-%m-%d')) \
            .order('Data de Disparo').range(offset, offset + page - 1).execute()
        batch = res.data or []
        rows.extend(batch)
        if len(batch) < page:
            break
    print(f'  Linhas do Supabase: {len(rows)}')
    return rows

# ── Indexação ─────────────────────────────────────────────────────────────────
# idx[tab][bloco][date_str][canal_col_rel] = {'sent': x, 'del': y, 'abe': z, 'clk': w}
def build_index(rows: list[dict], tabs_blocks: dict[str, list]) -> dict:
    idx = {}

    for row in rows:
        jornada = row.get('jornada') or ''
        canal_raw = row.get('Canal') or ''
        dt = parse_date(row.get('Data de Disparo'))
        if not dt:
            continue
        ds = dt.strftime('%Y-%m-%d')

        sent = int(row.get('Base Total') or 0)
        deliv = int(row.get('Base Acionável') or 0)
        abe   = int(row.get('Abertura') or 0)
        clk   = int(row.get('Cliques') or 0)

        # Normalizar canal
        c = canal_raw.upper()
        if 'WHATSAPP' in c or c == 'WPP': canal = 'WhatsApp'
        elif 'MAIL' in c:                  canal = 'E-mail'
        elif 'SMS' in c:                   canal = 'SMS'
        elif 'PUSH' in c:                  canal = 'Push'
        else:                              canal = None

        if canal is None:
            continue

        # Determinar em qual tab e bloco esta linha cai
        for tab_name, blocks in tabs_blocks.items():
            bloco = classify_bloco(jornada, blocks)
            if bloco:
                if tab_name not in idx:             idx[tab_name] = {}
                if bloco not in idx[tab_name]:      idx[tab_name][bloco] = {}
                if ds not in idx[tab_name][bloco]:  idx[tab_name][bloco][ds] = {}

                canal_map = CANAL_COLS.get(canal, {})
                for field, col_rel in canal_map.items():
                    val_map = {'sent': sent, 'del': deliv, 'abe': abe, 'clk': clk}
                    v = val_map.get(field, 0)
                    cur = idx[tab_name][bloco][ds].get(col_rel, 0)
                    idx[tab_name][bloco][ds][col_rel] = cur + v
                break  # first matching tab wins

    return idx

# ── Escrita da aba diarizada ──────────────────────────────────────────────────
def write_tab(wb, tab_name: str, blocks: list[tuple], idx_tab: dict, dates: list[datetime]):
    ws = wb.create_sheet(tab_name)
    ws.freeze_panes = 'C5'
    ws.sheet_view.showGridLines = False

    n_blocks = len(blocks)
    total_cols = FIXED_COLS + n_blocks * BLOCK_WIDTH

    # Linha 1 — KPI summary strip (placeholder, preenchível manualmente)
    ws.row_dimensions[1].height = 24
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=FIXED_COLS)
    c1 = cell(ws, 1, 1)
    style(c1, f'{tab_name} — Copa 2026', bold=True, fill=C['kpi_fill'],
          font_color='FFFFFF', align='left', size=11)
    # Totais por bloco na linha 1
    for bi, (bname, _) in enumerate(blocks):
        bc = FIXED_COLS + bi * BLOCK_WIDTH + 1
        ws.merge_cells(start_row=1, start_column=bc, end_row=1, end_column=bc + BLOCK_WIDTH - 1)
        style(cell(ws, 1, bc), bname.upper(), bold=True,
              fill=BLOCK_COLORS.get(bname, C['header_dark']),
              font_color='FFFFFF', size=10)

    # Linha 2 — nome dos blocos (span do bloco inteiro)
    ws.row_dimensions[2].height = 20
    for j in range(1, FIXED_COLS + 1):
        style(cell(ws, 2, j), fill=C['header_dark'], font_color='94A3B8')
    for bi, (bname, _) in enumerate(blocks):
        bc = FIXED_COLS + bi * BLOCK_WIDTH + 1
        for j in range(bc, bc + BLOCK_WIDTH):
            style(cell(ws, 2, j), fill=BLOCK_COLORS.get(bname, C['header_dark']))

    # Linha 3 — sub-header canal (wpp / e-mail / sms / push)
    ws.row_dimensions[3].height = 16
    fixed_labels_3 = ['', '', 'LP Tráfego', 'LP Optins', 'Mídia (R$)', 'Cliques', 'CPC']
    for j, lbl in enumerate(fixed_labels_3, 1):
        style(cell(ws, 3, j), lbl, fill=C['header_dark'], font_color='94A3B8', size=8)
    for bi, (bname, _) in enumerate(blocks):
        bc = FIXED_COLS + bi * BLOCK_WIDTH + 1
        for rel, lbl in enumerate(CANAL_SUBHEADER):
            c3 = cell(ws, 3, bc + rel)
            fill, font = canal_fill(rel)
            style(c3, lbl or '', fill=fill, font_color=font, bold=(lbl is not None), size=8)

    # Linha 4 — métricas
    ws.row_dimensions[4].height = 18
    fixed_labels_4 = ['Data', 'Dia', '', '', '', '', '']
    for j, lbl in enumerate(fixed_labels_4, 1):
        style(cell(ws, 4, j), lbl, bold=True, fill=C['header_dark'],
              font_color='FFFFFF', size=9)
    for bi in range(n_blocks):
        bc = FIXED_COLS + bi * BLOCK_WIDTH + 1
        for rel, lbl in enumerate(METRIC_SUBHEADER):
            c4 = cell(ws, 4, bc + rel)
            fill, font = canal_fill(rel)
            style(c4, lbl, bold=True, fill=fill, font_color=font, size=8)

    # Linhas de dados
    bloco_names = [b[0] for b in blocks]
    for di, d in enumerate(dates):
        row = 5 + di
        ws.row_dimensions[row].height = 14
        is_we = d.weekday() >= 5
        base_fill = C['weekend'] if is_we else (C['zebra'] if di % 2 else 'FFFFFF')

        # Data e Dia
        c_data = cell(ws, row, 1)
        c_data.value = d
        c_data.number_format = 'DD/MM'
        style(c_data, fill=base_fill, italic=is_we, size=9)

        style(cell(ws, row, 2), DAY_NAMES_PT[d.weekday()],
              fill=base_fill, italic=is_we, size=9)

        # Colunas fixas LP/Mídia (vazias por enquanto)
        for j in range(3, FIXED_COLS + 1):
            style(cell(ws, row, j), fill=base_fill, size=9)

        # Blocos CRM
        ds = d.strftime('%Y-%m-%d')
        for bi, bname in enumerate(bloco_names):
            bc = FIXED_COLS + bi * BLOCK_WIDTH + 1
            day_data = (idx_tab.get(bname) or {}).get(ds) or {}

            for rel in range(BLOCK_WIDTH):
                c_val = cell(ws, row, bc + rel)
                val = day_data.get(rel)
                fill, font = canal_fill(rel)
                if val:
                    style(c_val, val, fill=fill, font_color=font, size=9)
                    c_val.number_format = '#,##0'
                else:
                    style(c_val, fill=base_fill, size=9)

    # Linha de totais
    total_row = 5 + len(dates)
    ws.row_dimensions[total_row].height = 16
    style(cell(ws, total_row, 1), 'TOTAL', bold=True, fill=C['total_fill'], size=9)
    style(cell(ws, total_row, 2), '', fill=C['total_fill'], size=9)
    for j in range(3, FIXED_COLS + 1):
        style(cell(ws, total_row, j), fill=C['total_fill'])
    for bi in range(n_blocks):
        bc = FIXED_COLS + bi * BLOCK_WIDTH + 1
        for rel in range(BLOCK_WIDTH):
            c_tot = cell(ws, total_row, bc + rel)
            letter = get_column_letter(bc + rel)
            c_tot.value = f'=SUM({letter}5:{letter}{total_row-1})'
            c_tot.number_format = '#,##0'
            fill, font = canal_fill(rel)
            style(c_tot, fill=C['total_fill'], font_color=C['header_dark'], bold=True, size=9)
            c_tot.value = f'=SUM({letter}5:{letter}{total_row-1})'

    # Larguras
    ws.column_dimensions['A'].width = 9
    ws.column_dimensions['B'].width = 5
    ws.column_dimensions[get_column_letter(3)].width = 10
    ws.column_dimensions[get_column_letter(4)].width = 9
    ws.column_dimensions[get_column_letter(5)].width = 12
    ws.column_dimensions[get_column_letter(6)].width = 9
    ws.column_dimensions[get_column_letter(7)].width = 7
    for bi in range(n_blocks):
        bc = FIXED_COLS + bi * BLOCK_WIDTH + 1
        for rel in range(BLOCK_WIDTH):
            lbl = METRIC_SUBHEADER[rel]
            ws.column_dimensions[get_column_letter(bc + rel)].width = (
                11 if lbl == 'enviados' else 10 if lbl in ('entregas', 'abertura') else 9
            )

# ── Aba Auditoria ─────────────────────────────────────────────────────────────
def write_audit(wb, rows: list[dict], tabs_blocks: dict[str, list]):
    ws = wb.create_sheet('Auditoria')
    ws.sheet_view.showGridLines = False

    hdr_fill = C['header_dark']

    style(cell(ws, 1, 1), 'Auditoria — Rentabilização CRM',
          bold=True, fill=hdr_fill, font_color='FFFFFF', align='left', size=11)
    ws.merge_cells('A1:H1')

    # Resumo por aba/bloco
    summary: dict[str, dict[str, int]] = {}
    unmapped = []

    for row in rows:
        jornada = row.get('jornada') or ''
        matched = False
        for tab_name, blocks in tabs_blocks.items():
            bloco = classify_bloco(jornada, blocks)
            if bloco:
                summary.setdefault(tab_name, {})
                summary[tab_name][bloco] = summary[tab_name].get(bloco, 0) + 1
                matched = True
                break
        if not matched:
            unmapped.append(jornada)

    cursor = 3
    for tab_name, blocos in summary.items():
        style(cell(ws, cursor, 1), f'Aba: {tab_name}',
              bold=True, fill=hdr_fill, font_color='FFFFFF', align='left')
        ws.merge_cells(start_row=cursor, start_column=1, end_row=cursor, end_column=3)
        cursor += 1
        for bloco, qtd in sorted(blocos.items()):
            style(cell(ws, cursor, 1), bloco, align='left', size=9)
            style(cell(ws, cursor, 2), qtd, size=9)
            cursor += 1
        cursor += 1

    if unmapped:
        cursor += 1
        style(cell(ws, cursor, 1), 'Jornadas sem bloco mapeado',
              bold=True, fill='EF4444', font_color='FFFFFF', align='left')
        ws.merge_cells(start_row=cursor, start_column=1, end_row=cursor, end_column=4)
        cursor += 1
        by_jornada: dict[str, int] = {}
        for j in unmapped:
            by_jornada[j] = by_jornada.get(j, 0) + 1
        for j, qtd in sorted(by_jornada.items()):
            style(cell(ws, cursor, 1), j, align='left', size=8)
            style(cell(ws, cursor, 2), qtd, size=9)
            cursor += 1

    ws.column_dimensions['A'].width = 60
    ws.column_dimensions['B'].width = 10

# ── Main ───────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument('--start', default=None, help='Data início YYYY-MM-DD')
    parser.add_argument('--end',   default=None, help='Data fim YYYY-MM-DD')
    parser.add_argument('--output', default=None, help='Arquivo de saída .xlsx')
    args = parser.parse_args()

    now = datetime.now()
    start = datetime.strptime(args.start, '%Y-%m-%d') if args.start else datetime(now.year, now.month, 1)
    end   = datetime.strptime(args.end,   '%Y-%m-%d') if args.end   else datetime(now.year, now.month + 1 if now.month < 12 else 1,
                                                                                   1, now.year if now.month < 12 else now.year + 1) - timedelta(days=1)
    # corrigir se mês > 12
    if args.end is None:
        import calendar
        last = calendar.monthrange(now.year, now.month)[1]
        end = datetime(now.year, now.month, last)

    out_dir = Path('outputs')
    out_dir.mkdir(exist_ok=True)
    ts = now.strftime('%Y%m%d_%H%M%S')
    output = args.output or str(out_dir / f'rentabilizacao_crm_{ts}.xlsx')

    print(f'Periodo: {start.strftime("%d/%m/%Y")} ate {end.strftime("%d/%m/%Y")}')
    print(f'Output: {output}')

    # Fetch
    print('\nBuscando dados no Supabase...')
    rows = fetch_rows(start, end)

    if not rows:
        print('Nenhum dado encontrado para o período.')
        sys.exit(0)

    dates = date_range(start, end)

    # Tabs e blocos
    tabs_blocks = {
        'Rentabilização': RNT_BLOCKS,
        'Seguros':        SEG_BLOCKS,
    }

    # Index
    print('Indexando...')
    idx = build_index(rows, tabs_blocks)

    # Gerar Excel
    print('Gerando Excel...')
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove sheet padrão

    for tab_name, blocks in tabs_blocks.items():
        print(f'  Escrevendo aba: {tab_name}')
        write_tab(wb, tab_name, blocks, idx.get(tab_name, {}), dates)

    write_audit(wb, rows, tabs_blocks)

    wb.save(output)

    # Resumo
    total = len(rows)
    print(f'\nConcluído: {total} linhas | {len(dates)} dias | {output}')

    # Stats por aba/bloco
    for tab_name, blocks in tabs_blocks.items():
        print(f'\n  {tab_name}:')
        for bname, _ in blocks:
            bdata = (idx.get(tab_name) or {}).get(bname) or {}
            dias_com_dados = len([d for d in bdata.values() if any(v > 0 for v in d.values())])
            print(f'    {bname}: {dias_com_dados} dias com dados')

if __name__ == '__main__':
    main()
